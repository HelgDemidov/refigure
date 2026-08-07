"""API-level robustness tests: adversarial/edge-case input
through the public docx.convert()/xlsx.convert() surface, not the real
corpus-fixture behavioral tests (stage 5, gated on stage 3's licensing).

Covers: zip-bomb rejection through the public API, truncated/corrupted
zip, zero-byte input, non-ASCII round-trip, absence of shared state across
repeated calls, and (security-audit-remediation §1) ``zipsafe.safe_read``/
``zipsafe.check_archive``'s entry-count cap tested directly against
``refigure.core.zipsafe``, not just through the public API.
"""

from __future__ import annotations

import io
import struct
import time
import zipfile

import openpyxl
import pytest

import refigure.docx as docx
import refigure.xlsx as xlsx
from refigure.api import CorruptArchiveError, UnsupportedFormatError
from refigure.core import zipsafe

from .docx.test_docx import build_minimal_docx


def _oversized_member_zip() -> bytes:
    """A zip with one member whose declared uncompressed size (129MB) is
    just over zipsafe's 128MB per-member limit — highly compressible, so
    building it is fast and the actual bytes on disk stay tiny (~128KB)."""
    size = 129 * 1024 * 1024
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("word/document.xml", b"A" * size)
    return buf.getvalue()


def _oversized_total_zip() -> bytes:
    """5 members, each under zipsafe's 128MB per-member limit, whose
    declared uncompressed sizes SUM past its 512MB total-decompressed
    limit — the other zipsafe branch (aggregate, not any single member).
    Same highly-compressible-data trick as ``_oversized_member_zip``."""
    member_size = 110 * 1024 * 1024  # 5 * 110MB = 550MB > 512MB, each < 128MB
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for i in range(5):
            z.writestr(f"word/media/image{i}.bin", b"A" * member_size)
    return buf.getvalue()


def _corrupt_member_crc(data: bytes) -> bytes:
    """Flip a handful of bytes in the middle of a valid zip — corrupts a
    member's compressed data (bad CRC-32) without touching the central
    directory, so ZipFile() itself still opens fine."""
    mutable = bytearray(data)
    mid = len(mutable) // 2
    for i in range(mid, mid + 20):
        mutable[i] ^= 0xFF
    return bytes(mutable)


def _spoofed_declared_size_zip(
    real_size: int, spoofed_size: int, *, name: str = "word/document.xml"
) -> bytes:
    """A single-member zip whose LOCAL FILE HEADER and CENTRAL DIRECTORY
    "uncompressed size" fields are both patched to ``spoofed_size`` while
    the real (highly-compressible, cheap to build) payload still inflates
    to ``real_size`` bytes — ``ZipInfo.file_size`` (what ``check_archive()``
    alone trusts) lies about the real content, exactly the live audit PoC
    zipsafe.py's module docstring describes.

    ``spoofed_size`` must stay above ``zipsafe``'s internal 1MB read-chunk
    size for this to exercise ``safe_read()``'s own byte-counting check —
    if the declared size collapsed to 0 remaining quota within the FIRST
    chunk, zipfile's own (unrelated) CRC bookkeeping would raise
    ``BadZipFile`` before ``safe_read`` ever got a chance to compare
    against ``max_bytes``. Byte offsets: local file header's uncompressed
    size field is a 4-byte LE uint at offset 22 (after the 4-byte
    signature/2-byte version/2-byte flags/2-byte method/2-byte time/2-byte
    date/4-byte CRC-32/4-byte compressed-size fields); the central
    directory's own copy of the same field sits at offset 24 from that
    entry's ``PK\\x01\\x02`` signature (standard PKZIP APPNOTE layout)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr(name, b"A" * real_size)
    data = bytearray(buf.getvalue())
    assert data[0:4] == b"PK\x03\x04"
    struct.pack_into("<I", data, 22, spoofed_size)
    cd_offset = data.find(b"PK\x01\x02")
    assert cd_offset != -1
    struct.pack_into("<I", data, cd_offset + 24, spoofed_size)
    return bytes(data)


class TestZipBomb:
    def test_docx_convert_rejects_oversized_member(self) -> None:
        # Fixture construction (~1s, building the compressible 129MB member)
        # must not count against the rejection-speed budget below — time
        # only the convert() call itself.
        data = _oversized_member_zip()
        t0 = time.time()
        with pytest.raises(CorruptArchiveError):
            docx.convert(data)
        assert time.time() - t0 < 2.0  # rejected on declared size, never decompresses

    def test_xlsx_convert_rejects_oversized_member(self) -> None:
        data = _oversized_member_zip()
        t0 = time.time()
        with pytest.raises(CorruptArchiveError):
            xlsx.convert(data)
        assert time.time() - t0 < 2.0

    def test_docx_convert_rejects_oversized_total(self) -> None:
        data = _oversized_total_zip()
        t0 = time.time()
        with pytest.raises(CorruptArchiveError):
            docx.convert(data)
        assert time.time() - t0 < 2.0  # rejected on declared size, never decompresses

    def test_xlsx_convert_rejects_oversized_total(self) -> None:
        data = _oversized_total_zip()
        t0 = time.time()
        with pytest.raises(CorruptArchiveError):
            xlsx.convert(data)
        assert time.time() - t0 < 2.0


class TestSafeRead:
    """zipsafe.safe_read() itself (security-audit-remediation §1, findings
    #1/#2): check_archive()'s upfront gate trusts ZipInfo.file_size, which
    is attacker-controlled header metadata — safe_read() is the real
    enforcement, checked against ACTUAL decompressed bytes during the read.
    Tested directly against refigure.core.zipsafe, not just through the
    docx.convert()/xlsx.convert() public API surface used elsewhere in this
    file."""

    def test_rejects_spoofed_declared_size_during_the_read(self) -> None:
        # Declared size (patched into BOTH the local file header and the
        # central directory) understates the real payload by 2MB — a
        # spoofed-small declared size on a member whose real content is
        # much larger, the exact live-audit PoC. spoofed_size stays above
        # zipsafe's internal 1MB read-chunk size (see
        # _spoofed_declared_size_zip's docstring) so this exercises
        # safe_read's own byte-counting check, not zipfile's unrelated
        # CRC bookkeeping.
        data = _spoofed_declared_size_zip(real_size=5 * 1024 * 1024, spoofed_size=3 * 1024 * 1024)
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            info = z.getinfo("word/document.xml")
            assert info.file_size == 3 * 1024 * 1024  # the header patch took — declared size lies

            t0 = time.time()
            with pytest.raises(zipsafe.ArchiveBombSuspected):
                # max_bytes (500KB) is far smaller than both the spoofed
                # declared size (3MB) and the real payload (5MB) — a
                # correct implementation raises after the first ~1MB
                # chunk, well under 1s; one that buffered the whole real
                # payload first (or trusted the declared size) would not.
                zipsafe.safe_read(z, info, max_bytes=500_000)
            assert time.time() - t0 < 1.0

    def test_passes_through_an_honest_large_member_unchanged(self) -> None:
        # No false positive: an honestly-declared member within max_bytes
        # reads through safe_read cleanly and returns the exact real bytes.
        content = b"B" * (3 * 1024 * 1024)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
            z.writestr("word/document.xml", content)
        with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as z:
            result = zipsafe.safe_read(z, "word/document.xml", max_bytes=zipsafe.MAX_MEMBER_BYTES)
        assert result == content

    def test_accepts_either_a_member_name_or_a_zipinfo(self) -> None:
        # Same two accepted argument forms as zipfile.ZipFile.read().
        content = b"hello"
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.writestr("a.txt", content)
        with zipfile.ZipFile(io.BytesIO(buf.getvalue())) as z:
            assert zipsafe.safe_read(z, "a.txt") == content
            assert zipsafe.safe_read(z, z.getinfo("a.txt")) == content


class TestCorruptedZip:
    def test_docx_convert_rejects_truncated_zip(self) -> None:
        data = build_minimal_docx(["Hello"])
        truncated = data[: len(data) // 2]
        with pytest.raises(CorruptArchiveError):
            docx.convert(truncated)

    def test_xlsx_convert_rejects_truncated_zip(self) -> None:
        wb = openpyxl.Workbook()
        wb.active.append(["a"])
        buf = io.BytesIO()
        wb.save(buf)
        truncated = buf.getvalue()[: len(buf.getvalue()) // 2]
        with pytest.raises(CorruptArchiveError):
            xlsx.convert(truncated)

    def test_docx_convert_rejects_bad_crc(self) -> None:
        data = build_minimal_docx(["Hello"])
        with pytest.raises(CorruptArchiveError):
            docx.convert(_corrupt_member_crc(data))

    def test_xlsx_convert_rejects_bad_crc(self) -> None:
        wb = openpyxl.Workbook()
        wb.active.append(["a"])
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(CorruptArchiveError):
            xlsx.convert(_corrupt_member_crc(buf.getvalue()))


class TestZeroByteInput:
    def test_docx_convert_rejects_empty_bytes(self) -> None:
        with pytest.raises(CorruptArchiveError):
            docx.convert(b"")

    def test_xlsx_convert_rejects_empty_bytes(self) -> None:
        with pytest.raises(CorruptArchiveError):
            xlsx.convert(b"")


class TestNonAsciiRoundTrip:
    def test_docx_cyrillic_and_emoji_round_trip(self) -> None:
        text = "Привет, мир! 🎉 日本語"
        result = docx.convert(build_minimal_docx([text]))
        assert text in result.markdown

    def test_xlsx_cyrillic_and_emoji_round_trip(self) -> None:
        wb = openpyxl.Workbook()
        wb.active.append(["Привет", "🎉 日本語"])
        buf = io.BytesIO()
        wb.save(buf)
        result = xlsx.convert(buf.getvalue())
        assert "Привет" in result.markdown
        assert "🎉 日本語" in result.markdown


class TestNoSharedStateAcrossCalls:
    def test_repeated_docx_calls_are_independent(self) -> None:
        good = build_minimal_docx(["Stable content"])
        bad_format = _oversized_member_zip()

        # Interleave valid/invalid calls — a previous call's exception or
        # accumulated state must not leak into the next, unrelated call.
        for _ in range(5):
            result = docx.convert(good)
            assert result.markdown.strip().startswith("Stable content")
            assert result.warnings == []

            with pytest.raises(CorruptArchiveError):
                docx.convert(bad_format)

        # And once more, to confirm the good path still works cleanly after
        # N interleaved failures.
        final = docx.convert(good)
        assert final.markdown.strip().startswith("Stable content")
        assert final.warnings == []

    def test_repeated_xlsx_calls_are_independent(self) -> None:
        wb = openpyxl.Workbook()
        wb.active.append(["stable", "content"])
        buf = io.BytesIO()
        wb.save(buf)
        good = buf.getvalue()
        bad_format = _oversized_member_zip()

        for _ in range(5):
            result = xlsx.convert(good)
            assert "stable" in result.markdown
            assert result.warnings == []

            with pytest.raises(CorruptArchiveError):
                xlsx.convert(bad_format)

        final = xlsx.convert(good)
        assert "stable" in final.markdown
        assert final.warnings == []

    def test_docx_and_xlsx_calls_do_not_interfere_with_each_other(self) -> None:
        docx_data = build_minimal_docx(["From docx"])
        wb = openpyxl.Workbook()
        wb.active.append(["From xlsx"])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_data = buf.getvalue()

        for _ in range(3):
            with pytest.raises(UnsupportedFormatError):
                docx.convert(xlsx_data)  # valid xlsx is not a valid docx
            with pytest.raises(UnsupportedFormatError):
                xlsx.convert(docx_data)  # and vice versa

        assert "From docx" in docx.convert(docx_data).markdown
        assert "From xlsx" in xlsx.convert(xlsx_data).markdown
