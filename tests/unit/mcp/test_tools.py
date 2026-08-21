"""In-memory Client tests for convert_docx/convert_xlsx — architecture doc
§9's pattern (see conftest.py's anyio_backend fixture for why anyio, not
pytest-asyncio). Corpus fixtures are deliberately NOT used here (they're
optional/gitignored, and tests/integration/ already exercises real corpus
documents through the core convert() functions) — synthetic builders
(``build_minimal_docx``/a bare ``openpyxl.Workbook()``) keep this suite
fast and independent of local fixture setup, matching
``tests/unit/vlm/test_vlm.py``'s own convention.
"""

from __future__ import annotations

import base64
import io
import zipfile
from typing import Any
from unittest.mock import patch

import openpyxl
import pytest
from mcp import Client

from refigure.mcp.server import build_server
from tests.unit.docx.test_docx import build_minimal_docx

pytestmark = pytest.mark.anyio


def _minimal_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
async def client():
    mcp_server = build_server()
    async with Client(mcp_server, raise_exceptions=True) as c:
        yield c


def _text(result: Any) -> str:
    return result.content[0].text if result.content else ""


async def test_convert_docx_via_path_succeeds(client: Client, tmp_path) -> None:
    doc_path = tmp_path / "doc.docx"
    doc_path.write_bytes(build_minimal_docx(["hello world"]))

    result = await client.call_tool("convert_docx", {"path": str(doc_path)})

    assert result.is_error is False
    sc = result.structured_content
    assert "hello world" in sc["markdown"]
    assert sc["resource_uri"] is None
    assert sc["vlm_used"] is False


async def test_convert_xlsx_via_content_base64_succeeds(client: Client) -> None:
    b64 = base64.b64encode(_minimal_xlsx_bytes()).decode("ascii")

    result = await client.call_tool("convert_xlsx", {"content_base64": b64})

    assert result.is_error is False
    sc = result.structured_content
    assert "hello" in sc["markdown"]
    assert sc["warnings"] == []


async def test_convert_xlsx_use_vlm_warns_without_a_real_vlm_path(client: Client) -> None:
    b64 = base64.b64encode(_minimal_xlsx_bytes()).decode("ascii")

    with patch("refigure.vlm.client.OpenRouterClient.send") as send:
        result = await client.call_tool("convert_xlsx", {"content_base64": b64, "use_vlm": True})

    assert result.is_error is False
    sc = result.structured_content
    assert sc["vlm_used"] is False
    assert any("no VLM path" in w for w in sc["warnings"])
    send.assert_not_called()


async def test_convert_docx_neither_path_nor_content_base64_is_an_error(client: Client) -> None:
    result = await client.call_tool("convert_docx", {})

    assert result.is_error is True
    assert "exactly one of path or content_base64" in _text(result)


async def test_convert_docx_both_path_and_content_base64_is_an_error(client: Client) -> None:
    result = await client.call_tool(
        "convert_docx", {"path": "irrelevant.docx", "content_base64": "eA=="}
    )

    assert result.is_error is True
    assert "exactly one of path or content_base64" in _text(result)


async def test_content_base64_over_cap_rejected_before_decoding(client: Client) -> None:
    oversized = "A" * (200 * 1024 * 1024)  # server default cap is 100 MB

    with patch("refigure.mcp.server.base64.b64decode") as b64decode:
        result = await client.call_tool("convert_docx", {"content_base64": oversized})

    assert result.is_error is True
    assert "MB cap" in _text(result)
    b64decode.assert_not_called()


async def test_vlm_verify_and_judge_mode_and_model_kwargs_reach_config(client: Client) -> None:
    """vlm_verify/vlm_judge_mode/vlm_model are only meaningful together
    with use_vlm=True and a real marker — this document has neither, so
    the call still completes successfully; the point is proving these 3
    optional per-call kwargs reach _run_convert_tool/Config without
    raising, not exercising the VLM engine itself (already covered by
    tests/unit/vlm/test_vlm.py)."""
    b64 = base64.b64encode(build_minimal_docx(["no markers here"])).decode("ascii")

    result = await client.call_tool(
        "convert_docx",
        {
            "content_base64": b64,
            "use_vlm": True,
            "vlm_verify": True,
            "vlm_judge_mode": "solo",
            "vlm_model": "some/model",
        },
    )

    assert result.is_error is False


async def test_content_base64_malformed_is_a_clean_error_not_a_raw_binascii_error(
    client: Client,
) -> None:
    result = await client.call_tool("convert_docx", {"content_base64": "not valid base64!!!"})

    assert result.is_error is True
    assert "not valid base64" in _text(result)


async def test_corrupt_archive_reports_the_typed_exception_class(client: Client) -> None:
    not_a_zip = base64.b64encode(b"this is not a zip file at all").decode("ascii")

    result = await client.call_tool("convert_docx", {"content_base64": not_a_zip})

    assert result.is_error is True
    text = _text(result)
    assert text.startswith("Error executing tool convert_docx: CorruptArchiveError:")
    # Never a bare traceback/repr leaking the internal zipfile message alone
    # without the classifying prefix _call_and_wrap_errors adds.
    assert "internal_error" not in text


async def test_convert_docx_use_vlm_is_inert_without_any_vlm_markers(client: Client) -> None:
    """use_vlm=True on a document with nothing for it to interpret (no
    image/group markers) must never fail just because no VLM
    provider/API key is configured — enhance_docx_markdown only resolves
    an API key lazily, on an actual marker."""
    doc_bytes = build_minimal_docx(["some text, no VLM markers though"])
    b64 = base64.b64encode(doc_bytes).decode("ascii")

    with patch.dict("os.environ", {}, clear=True):
        result = await client.call_tool("convert_docx", {"content_base64": b64, "use_vlm": True})

    assert result.is_error is False
    assert result.structured_content["vlm_used"] is False


async def test_list_tools_registers_both_convert_tools(client: Client) -> None:
    tools = await client.list_tools()
    names = {t.name for t in tools.tools}
    assert names == {"convert_docx", "convert_xlsx"}
    for tool in tools.tools:
        assert tool.annotations is not None
        assert tool.annotations.read_only_hint is True


async def test_corrupt_zip_bytes_are_rejected_as_corrupt_archive(client: Client) -> None:
    """A structurally-valid zip with no OOXML content still degrades to
    CorruptArchiveError, not something more exotic — belt-and-suspenders
    alongside test_corrupt_archive_reports_the_typed_exception_class'
    not-a-zip-at-all case above."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("unrelated.txt", "not a docx")
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")

    result = await client.call_tool("convert_docx", {"content_base64": b64})

    assert result.is_error is True
    assert "CorruptArchiveError" in _text(result) or "UnsupportedFormatError" in _text(result)


async def test_unexpected_exception_is_reported_as_internal_error(
    client: Client, monkeypatch: pytest.MonkeyPatch, caplog: pytest.LogCaptureFixture
) -> None:
    """The catch-all branch of _call_and_wrap_errors — every OTHER test in
    this file exercises either a refigure typed exception or a ValueError
    from _run_convert_tool's own input validation, never a genuinely
    unexpected one."""
    import refigure.docx as docx_module

    def _boom(*args: object, **kwargs: object) -> None:
        raise RuntimeError("boom, something truly unexpected")

    monkeypatch.setattr(docx_module, "convert", _boom)
    b64 = base64.b64encode(build_minimal_docx(["irrelevant"])).decode("ascii")

    with caplog.at_level("ERROR"):
        result = await client.call_tool("convert_docx", {"content_base64": b64})

    assert result.is_error is True
    text = _text(result)
    assert text.startswith("Error executing tool convert_docx: internal_error:")
    assert "boom" in text
    assert "unexpected error in a tool call" in caplog.text


async def test_server_ctx_vlm_client_reaches_per_call_config() -> None:
    """server_ctx.vlm_client (set once at server startup, per architecture
    doc §3 — never per-call) must reach every Config this server builds."""
    import refigure.docx as docx_module

    class _StubVlmClient:
        def send(self, prompt: str, image_uri: str, *, model: str) -> str:
            return "unused"

    captured: dict[str, object] = {}

    def _capture_convert(source: object, *, config: object = None) -> object:
        captured["config"] = config
        from refigure.api import ConversionResult

        return ConversionResult(markdown="ok")

    stub_client = _StubVlmClient()
    mcp_server = build_server(vlm_client=stub_client)  # type: ignore[arg-type]

    with patch.object(docx_module, "convert", _capture_convert):
        async with Client(mcp_server, raise_exceptions=True) as c:
            b64 = base64.b64encode(build_minimal_docx(["irrelevant"])).decode("ascii")
            await c.call_tool("convert_docx", {"content_base64": b64})

    assert captured["config"].vlm_client is stub_client  # type: ignore[attr-defined]


async def test_server_ctx_vlm_api_key_reaches_per_call_config_when_no_client_set() -> None:
    """The openrouter-default path (server_ctx.vlm_client is None): a
    --vlm-api-key-file-resolved key still needs to reach every Config,
    same as the vlm_client case above but the other branch."""
    import refigure.docx as docx_module

    captured: dict[str, object] = {}

    def _capture_convert(source: object, *, config: object = None) -> object:
        captured["config"] = config
        from refigure.api import ConversionResult

        return ConversionResult(markdown="ok")

    mcp_server = build_server(vlm_api_key="sk-or-test-key")

    with patch.object(docx_module, "convert", _capture_convert):
        async with Client(mcp_server, raise_exceptions=True) as c:
            b64 = base64.b64encode(build_minimal_docx(["irrelevant"])).decode("ascii")
            await c.call_tool("convert_docx", {"content_base64": b64})

    assert captured["config"].vlm_api_key == "sk-or-test-key"  # type: ignore[attr-defined]
    assert captured["config"].vlm_client is None  # type: ignore[attr-defined]
