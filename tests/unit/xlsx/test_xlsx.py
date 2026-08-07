"""Synthetic smoke tests for refigure.xlsx.convert().

Fixtures are built with openpyxl.Workbook() directly (already a hard
dependency of this module) — not the real corpus-fixture behavioral tests
(stage 5, gated on fixture licensing, stage 3), just enough to confirm the
new wrapper (stage 2) actually works end to end.
"""

from __future__ import annotations

import io
import zipfile

import openpyxl
import pytest
from openpyxl.chart import BarChart, Reference

from refigure import CorruptArchiveError, UnsupportedFormatError
from refigure.core import chart_render
from refigure.xlsx import convert


def _save(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_convert_returns_markdown_for_simple_workbook() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["col1", "col2"])
    ws.append([1, 2.5])

    result = convert(_save(wb))
    assert "## Sheet1" in result.markdown
    assert "col1" in result.markdown and "2.5" in result.markdown
    assert result.warnings == []
    assert (result.charts_found, result.charts_rendered, result.groups_found) == (0, 0, 0)


def test_convert_accepts_path_bytes_and_file_like_identically(tmp_path) -> None:
    wb = openpyxl.Workbook()
    wb.active.append(["a", "b"])
    data = _save(wb)
    path = tmp_path / "book.xlsx"
    path.write_bytes(data)

    from_path = convert(path)
    from_bytes = convert(data)
    from_stream = convert(io.BytesIO(data))

    assert from_path.markdown == from_bytes.markdown == from_stream.markdown


def test_empty_sheet_is_skipped_not_an_exception() -> None:
    # A workbook with only empty sheets IS "no content" (see the next test) —
    # this exercises the per-sheet skip marker alongside a sheet that has data.
    wb = openpyxl.Workbook()
    wb.active.title = "Data"
    wb.active.append(["x"])
    wb.create_sheet("Empty")

    result = convert(_save(wb))
    assert '[Sheet "Empty" — empty, skipped]' in result.markdown
    assert result.warnings == []


def test_all_sheets_empty_is_a_warning_not_an_exception() -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "Empty"
    wb.create_sheet("AlsoEmpty")

    result = convert(_save(wb))
    assert "no extractable content" in result.warnings


def test_chart_without_numcache_is_found_but_not_rendered() -> None:
    # openpyxl-authored charts carry formula references, not cached values —
    # the documented v1 zero-loss floor (caption-only marker), not a bug.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Category", "Value"])
    ws.append(["A", 10])
    ws.append(["B", 20])
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=3))
    ws.add_chart(chart, "D2")

    result = convert(_save(wb))
    assert result.charts_found == 1
    assert result.charts_rendered == 0
    assert "chart content not analyzed" in result.markdown


def _force_float_cell_xml(data: bytes) -> bytes:
    """openpyxl's own writer normalizes an integer-valued float like 5.0 to
    the bare text "5" (round-trips back as a Python int, not float — no
    decimal point in the cell XML for openpyxl's reader to key off of) —
    there's no way to get a genuine is_integer() float through openpyxl's
    write path alone. Patches the raw cell XML after the fact instead."""
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data)) as z, zipfile.ZipFile(buf, "w") as zo:
        for n in z.namelist():
            content = z.read(n)
            if n == "xl/worksheets/sheet1.xml":
                content = content.replace(b"<v>5</v>", b"<v>5.0</v>")
            zo.writestr(n, content)
    return buf.getvalue()


def test_integer_valued_float_cell_drops_trailing_dot_zero() -> None:
    # test_convert_returns_markdown_for_simple_workbook already covers the
    # non-integer float case (2.5) — this is the OTHER _xlsx_cell_str
    # branch, str(int(v)) instead of str(v).
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([5.0])
    data = _force_float_cell_xml(_save(wb))

    result = convert(data)

    assert "| 5 |" in result.markdown
    assert "5.0" not in result.markdown


def test_convert_warns_when_mermaidx_unavailable(monkeypatch: pytest.MonkeyPatch) -> None:
    # Same technique as test_chart_render_missing_mermaidx.py / the docx
    # side of this test (test_docx.py) — monkeypatch the module-level
    # optional import. charts_found increments regardless of whether the
    # chart itself renders (see test_chart_without_numcache_is_found_but_
    # not_rendered), so the no-numCache fixture is enough here.
    monkeypatch.setattr(chart_render, "mermaidx", None)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Category", "Value"])
    ws.append(["A", 10])
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=2), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=2))
    ws.add_chart(chart, "D2")

    result = convert(_save(wb))

    assert result.charts_found == 1
    assert any(w.startswith("mermaidx not installed") for w in result.warnings)


def test_valid_zip_but_not_xlsx_raises_unsupported_format() -> None:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("hello.txt", "hi")
    with pytest.raises(UnsupportedFormatError):
        convert(buf.getvalue())


def test_non_zip_raises_corrupt_archive() -> None:
    with pytest.raises(CorruptArchiveError):
        convert(b"not a zip at all")


def test_convert_still_works_for_an_ordinary_path(tmp_path) -> None:
    # Regression check for the normalize_source is_file() gate added for
    # security-audit finding #17: a genuine regular-file Path must keep
    # working, not just get rejected less broadly than intended.
    wb = openpyxl.Workbook()
    wb.active.append(["a", "b"])
    path = tmp_path / "book.xlsx"
    path.write_bytes(_save(wb))
    result = convert(path)
    assert "a" in result.markdown


def test_convert_rejects_a_fifo_path_promptly_instead_of_hanging(tmp_path) -> None:
    # Security-audit finding #17: a Path pointing at a FIFO/named pipe with
    # no writer used to make zipfile.ZipFile(path) (inside
    # zipsafe.check_archive) hang forever, reachable via the direct library
    # API (unlike the CLI, which already gates on is_file() before ever
    # reaching convert()). normalize_source's is_file() check must reject
    # it up front without ever opening it — verified here with an external
    # timeout as a safety net in case that regresses.
    import os

    from tests.unit.test_io import run_with_timeout

    fifo_path = tmp_path / "pipe"
    os.mkfifo(fifo_path)

    with pytest.raises(CorruptArchiveError):
        run_with_timeout(lambda: convert(fifo_path))
