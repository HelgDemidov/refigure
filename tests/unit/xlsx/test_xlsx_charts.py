"""Tests for xlsx_charts.py (spec convert-xlsx §3 + chart-data-extraction
§4.3): embedded-chart detection, captions from c:title, id12 derived from
the chart's XML structure. No network, no LibreOffice — pure in-memory XML
(openpyxl.chart builds a real chart part). Mutual-ownership/render-isolation
(``extract_chart_workbook`` and all print-area geometry) were removed
together with the VLM path — see ``chart_data.py``/``chart_render.py`` for
the data-driven replacement."""

from __future__ import annotations

import io
import zipfile
from pathlib import Path
from typing import Any

import pytest
from lxml import etree
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

import refigure.xlsx.charts as charts_module
from refigure.xlsx.charts import (
    _chart_anchors,
    _chart_refs,
    _filter_caption_texts,
    _q,
    _rel_targets,
    _resolve_target,
    _sheet_parts,
    extract_charts,
)


def _active(wb: Any) -> Any:
    """``wb.active`` is typed as ``Worksheet | Chartsheet | None`` in the
    stubs — a fresh ``Workbook()`` always carries an active Worksheet."""
    ws = wb.active
    assert ws is not None
    return ws


def _workbook_with_chart(
    tmp_path: Path,
    *,
    title: str | None = "Chart Title",
    anchor: str = "D2",
    sheet_name: str = "Data",
    file_name: str = "raw.xlsx",
) -> Path:
    wb = Workbook()
    ws = _active(wb)
    ws.title = sheet_name
    ws.append(["Cat", "Val"])
    ws.append(["A", 1])
    ws.append(["B", 2])
    chart = BarChart()
    if title is not None:
        chart.title = title
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)
    raw = tmp_path / file_name
    wb.save(raw)
    return raw


def test_extract_charts_empty_workbook_returns_empty_list(tmp_path: Path) -> None:
    wb = Workbook()
    raw = tmp_path / "raw.xlsx"
    wb.save(raw)
    assert extract_charts(raw) == []


def test_extract_charts_detects_single_chart_with_title_and_anchor(tmp_path: Path) -> None:
    raw = _workbook_with_chart(
        tmp_path, title="Costs of LTE and 5G", anchor="D2", sheet_name="Data"
    )
    charts = extract_charts(raw)
    assert len(charts) == 1
    chart = charts[0]
    assert chart.sheet == "Data"
    assert chart.anchor_cell == "D2"
    assert chart.captions == ("Costs of LTE and 5G",)


def test_extract_charts_no_title_gives_empty_captions(tmp_path: Path) -> None:
    raw = _workbook_with_chart(tmp_path, title=None)
    charts = extract_charts(raw)
    assert len(charts) == 1
    assert charts[0].captions == ()


def test_extract_charts_id12_stable_across_repeated_calls(tmp_path: Path) -> None:
    raw = _workbook_with_chart(tmp_path)
    id1 = extract_charts(raw)[0].id12
    id2 = extract_charts(raw)[0].id12
    assert id1 == id2
    assert len(id1) == 12


def test_extract_charts_different_titles_get_distinct_ids(tmp_path: Path) -> None:
    """The builder is designed for one chart per document (symmetric with the
    docx tests) — compare across two separate single-chart documents that
    id12 depends on the chart's content (the title is part of the chart
    part)."""
    raw_a = _workbook_with_chart(tmp_path, title="Title A", file_name="a.xlsx")
    raw_b = _workbook_with_chart(tmp_path, title="Title B", file_name="b.xlsx")
    id_a = extract_charts(raw_a)[0].id12
    id_b = extract_charts(raw_b)[0].id12
    assert id_a != id_b


def test_extract_charts_anchor_cell_reflects_position(tmp_path: Path) -> None:
    raw = _workbook_with_chart(tmp_path, anchor="G10")
    charts = extract_charts(raw)
    assert charts[0].anchor_cell == "G10"


# --- _chart_anchors (a pure function — hand-assembled XML, without openpyxl,
# structure mirrors a real fixture exactly, see below) ---


def test_chart_anchors_yields_every_chart_in_a_grouped_anchor() -> None:
    """Live bug (stage 5 corpus testing, 2026-08-05, root-caused via
    superpowers:systematic-debugging): a single xdr:twoCellAnchor can
    legitimately be an xdr:grpSp (Excel's "group these charts" feature)
    wrapping MULTIPLE charts at one shared position — confirmed real
    (eia-aeo-2026-figures.xlsx's drawing25.xml, one anchor nesting 6
    charts, each in its own graphicFrame; eia-ieo-2023-figures.xlsx and
    waste-statistics.xlsx hit the same shape). _chart_anchors used to use a
    singular .find() for c:chart, silently dropping every chart after the
    first in such a group — no warning, no trace anywhere. The structure
    below (wsDr > twoCellAnchor > grpSp > graphicFrame > graphic >
    graphicData > chart, x2) mirrors the real fixture's structure exactly,
    confirmed via direct inspection of drawing25.xml."""
    xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    c = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    a = "http://schemas.openxmlformats.org/drawingml/2006/main"
    r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    def graphic_frame(rid: str) -> str:
        return (
            "<xdr:graphicFrame><xdr:nvGraphicFramePr/><xdr:xfrm/>"
            f'<a:graphic><a:graphicData uri="{c}">'
            f'<c:chart r:id="{rid}"/></a:graphicData></a:graphic></xdr:graphicFrame>'
        )

    drawing_xml = (
        f'<xdr:wsDr xmlns:xdr="{xdr}" xmlns:c="{c}" xmlns:a="{a}" xmlns:r="{r}">'
        "<xdr:twoCellAnchor>"
        "<xdr:from><xdr:col>0</xdr:col><xdr:row>0</xdr:row></xdr:from>"
        "<xdr:grpSp><xdr:nvGrpSpPr/><xdr:grpSpPr/>"
        f"{graphic_frame('rId1')}{graphic_frame('rId2')}"
        "</xdr:grpSp></xdr:twoCellAnchor></xdr:wsDr>"
    )
    drawing_root = etree.fromstring(drawing_xml.encode())
    pairs = _chart_anchors(drawing_root)
    assert len(pairs) == 2
    rids = {chart_ref.get(_q("r", "id")) for _anchor, chart_ref in pairs}
    assert rids == {"rId1", "rId2"}


def test_chart_anchors_single_chart_anchor_still_yields_one() -> None:
    """Regression guard for the common case (no grouping) — the fix must not
    turn a normal, single-chart anchor into anything but one pair."""
    xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    c = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    a = "http://schemas.openxmlformats.org/drawingml/2006/main"
    r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    drawing_xml = (
        f'<xdr:wsDr xmlns:xdr="{xdr}" xmlns:c="{c}" xmlns:a="{a}" xmlns:r="{r}">'
        "<xdr:oneCellAnchor>"
        "<xdr:from><xdr:col>0</xdr:col><xdr:row>0</xdr:row></xdr:from>"
        "<xdr:graphicFrame><xdr:nvGraphicFramePr/><xdr:xfrm/>"
        f'<a:graphic><a:graphicData uri="{c}"><c:chart r:id="rId1"/>'
        "</a:graphicData></a:graphic></xdr:graphicFrame>"
        "</xdr:oneCellAnchor></xdr:wsDr>"
    )
    drawing_root = etree.fromstring(drawing_xml.encode())
    pairs = _chart_anchors(drawing_root)
    assert len(pairs) == 1
    assert pairs[0][1].get(_q("r", "id")) == "rId1"


# --- _filter_caption_texts: numeric junk / dedup / blank strings (a pure function) ---


def test_filter_caption_texts_skips_numeric_junk_duplicates_and_blanks() -> None:
    assert _filter_caption_texts(["Title", "42", "-3", "Title", "  ", "Subtitle"]) == (
        "Title",
        "Subtitle",
    )


# --- _rel_targets / _resolve_target (pure functions, malformed-chain cases) ---


def test_rel_targets_missing_rels_file_returns_empty(tmp_path: Path) -> None:
    """A fresh single-sheet workbook with no charts/hyperlinks doesn't carry
    xl/worksheets/_rels/sheet1.xml.rels at all."""
    raw = tmp_path / "raw.xlsx"
    Workbook().save(raw)
    with zipfile.ZipFile(raw) as z:
        assert _rel_targets(z, "xl/worksheets/sheet1.xml") == {}


def test_resolve_target_absolute_path_strips_leading_slash() -> None:
    resolved = _resolve_target("xl/worksheets/sheet1.xml", "/xl/media/image1.png")
    assert resolved == "xl/media/image1.png"


def test_resolve_target_relative_path_joins_with_source_dir() -> None:
    """openpyxl itself always writes ABSOLUTE targets (a live trace confirmed:
    '/xl/worksheets/sheet1.xml' etc.) — this branch is never hit by real
    openpyxl fixtures, but it's legitimate per the OPC spec (other
    writers/hand-assembled OOXML use relative targets) — a direct unit test
    on the pure function."""
    resolved = _resolve_target("xl/worksheets/sheet1.xml", "../drawings/drawing1.xml")
    assert resolved == "xl/drawings/drawing1.xml"


# --- _sheet_parts (pure, synthetic in-memory zip without openpyxl — severely
# malformed workbook.xml that no real writer would produce, but the parser
# must survive) ---


def test_sheet_parts_missing_workbook_xml_returns_empty() -> None:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w"):
        pass
    buf.seek(0)
    with zipfile.ZipFile(buf) as z:
        assert _sheet_parts(z) == {}


def test_sheet_parts_skips_sheet_element_without_rid() -> None:
    wb_root = etree.Element(_q("main", "workbook"))
    sheets_el = etree.SubElement(wb_root, _q("main", "sheets"))
    etree.SubElement(sheets_el, _q("main", "sheet")).set("name", "Data")  # no r:id
    wb_xml = etree.tostring(wb_root, xml_declaration=True, encoding="UTF-8", standalone=True)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("xl/workbook.xml", wb_xml)
    buf.seek(0)
    with zipfile.ZipFile(buf) as z:
        assert _sheet_parts(z) == {}


# --- extract_charts: a malformed/unreachable reference at each step of the
# OOXML chain (terminal safety net for a corrupted file, see extract_charts'
# docstring) — real openpyxl output with one link of the chain deliberately
# removed, not an artificially simplified XML ---


def _remove_zip_part(raw: Path, part_name: str) -> None:
    """In-place: rebuild the zip WITHOUT the given part — simulates a
    broken/unreachable reference in the OOXML chain on real openpyxl
    output."""
    orig = raw.read_bytes()
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(orig)) as z, zipfile.ZipFile(buf, "w") as zo:
        for n in z.namelist():
            if n != part_name:
                zo.writestr(n, z.read(n))
    raw.write_bytes(buf.getvalue())


def _find_part(raw: Path, *, startswith: str, endswith: str = ".xml") -> str:
    with zipfile.ZipFile(raw) as z:
        return next(n for n in z.namelist() if n.startswith(startswith) and n.endswith(endswith))


def test_extract_charts_skips_when_sheet_part_missing(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """iter_chart_entries' own ``sheet_part not in names`` check
    (independent of ``_sheet_parts``'s identical check at its own level,
    which already filters this case out before it ever reaches
    iter_chart_entries — the two can't diverge through a real zip, both
    read the same ``z.namelist()``) — exercised directly via a stubbed
    ``_sheet_parts`` that names a part absent from the real archive."""
    raw = _workbook_with_chart(tmp_path)
    monkeypatch.setattr(
        charts_module, "_sheet_parts", lambda z: {"Data": "xl/worksheets/does-not-exist.xml"}
    )
    assert extract_charts(raw) == []


def test_extract_charts_skips_when_sheet_rels_missing(tmp_path: Path) -> None:
    """The drawing r:id is present in the sheet itself, but the sheet's
    .rels is entirely missing -> drid not in {}."""
    raw = _workbook_with_chart(tmp_path)
    sheet_rels = _find_part(raw, startswith="xl/worksheets/_rels/", endswith=".rels")
    _remove_zip_part(raw, sheet_rels)
    assert extract_charts(raw) == []


def test_extract_charts_skips_when_drawing_part_missing(tmp_path: Path) -> None:
    """The sheet's rels correctly references the drawing, but
    xl/drawings/drawingN.xml itself is removed."""
    raw = _workbook_with_chart(tmp_path)
    drawing_part = _find_part(raw, startswith="xl/drawings/", endswith=".xml")
    _remove_zip_part(raw, drawing_part)
    assert extract_charts(raw) == []


def test_extract_charts_skips_when_drawing_rels_missing(tmp_path: Path) -> None:
    """The drawing references the chart part, but the drawing's own .rels
    is removed entirely."""
    raw = _workbook_with_chart(tmp_path)
    drawing_rels = _find_part(raw, startswith="xl/drawings/_rels/", endswith=".rels")
    _remove_zip_part(raw, drawing_rels)
    assert extract_charts(raw) == []


def test_extract_charts_skips_when_chart_part_missing(tmp_path: Path) -> None:
    """The drawing rels correctly references the chart, but
    xl/charts/chartN.xml itself is removed."""
    raw = _workbook_with_chart(tmp_path)
    chart_part = _find_part(raw, startswith="xl/charts/", endswith=".xml")
    _remove_zip_part(raw, chart_part)
    assert extract_charts(raw) == []


# --- _chart_refs (a pure function — hand-assembled XML, without openpyxl) ---


def test_chart_refs_skips_formula_without_sheet_qualifier() -> None:
    root = etree.Element(_q("c", "chartSpace"))
    etree.SubElement(root, _q("c", "f")).text = "A1"  # no "!" — not our format
    assert _chart_refs(root) == []


def test_chart_refs_unquotes_sheet_name_wrapped_in_apostrophes() -> None:
    root = etree.Element(_q("c", "chartSpace"))
    etree.SubElement(root, _q("c", "f")).text = "'My Sheet'!$A$1"
    assert _chart_refs(root) == [("My Sheet", "$A$1")]


def test_chart_refs_unescapes_doubled_apostrophe_within_quoted_sheet_name() -> None:
    # Standard Excel escaping: a literal apostrophe in a sheet name is
    # doubled inside the quoted form.
    root = etree.Element(_q("c", "chartSpace"))
    etree.SubElement(root, _q("c", "f")).text = "'O''Brien''s Sheet'!$A$1"
    assert _chart_refs(root) == [("O'Brien's Sheet", "$A$1")]


def test_chart_refs_skips_when_sheet_name_is_empty_after_unquoting() -> None:
    root = etree.Element(_q("c", "chartSpace"))
    etree.SubElement(root, _q("c", "f")).text = "''!$A$1"
    assert _chart_refs(root) == []


def test_extract_charts_sheet_with_no_drawing_returns_no_charts(tmp_path: Path) -> None:
    wb = Workbook()
    ws = _active(wb)
    ws.append(["plain", "data"])
    raw = tmp_path / "raw.xlsx"
    wb.save(raw)
    assert extract_charts(raw) == []
