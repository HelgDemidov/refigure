"""``tests/unit/test_cli.py`` — ``refigure.cli``, the argparse-based console
entry point. Synthetic docx/xlsx fixtures (``build_minimal_docx`` / a plain
``openpyxl.Workbook()``) for speed, same convention as
``test_robustness.py`` — real-corpus, end-to-end coverage lives in the
integration smoke test (full 27-fixture batch)."""

from __future__ import annotations

import io
import json
import subprocess
import sys
import zipfile
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

import refigure.cli as cli_module
from refigure.api import ConversionResult
from refigure.cli import (
    EXIT_BATCH_PARTIAL_FAILURE,
    EXIT_CORRUPT_ARCHIVE,
    EXIT_INTERNAL_ERROR,
    EXIT_MISSING_DEPENDENCY,
    EXIT_OK,
    EXIT_UNSUPPORTED_FORMAT,
    EXIT_USAGE,
    EXIT_VLM_MARKER_LIMIT,
    _plan_batch,
    main,
)
from tests.support import REPO_ROOT

from .docx.test_docx import build_minimal_docx


def _write_docx(path: Path, paragraphs: list[str]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(build_minimal_docx(paragraphs))
    return path


def _write_xlsx(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    assert wb.active is not None
    wb.active.append(["a", "b"])
    buf = io.BytesIO()
    wb.save(buf)
    path.write_bytes(buf.getvalue())
    return path


class TestSingleFile:
    def test_stdout_by_default(self, tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello world"])
        assert main([str(doc)]) == EXIT_OK
        assert "Hello world" in capsys.readouterr().out

    def test_output_file_not_stdout(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello world"])
        out_file = tmp_path / "out.md"
        assert main([str(doc), "-o", str(out_file)]) == EXIT_OK
        assert "Hello world" in out_file.read_text(encoding="utf-8")
        assert capsys.readouterr().out == ""

    def test_missing_file_is_usage_error(self, tmp_path: Path) -> None:
        with pytest.raises(SystemExit) as exc:
            main([str(tmp_path / "nope.docx")])
        assert exc.value.code == EXIT_USAGE

    def test_unrecognized_extension_is_usage_error(self, tmp_path: Path) -> None:
        stray = tmp_path / "doc.txt"
        stray.write_text("hi", encoding="utf-8")
        with pytest.raises(SystemExit) as exc:
            main([str(stray)])
        assert exc.value.code == EXIT_USAGE

    def test_xlsx_single_file_via_cli(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        # No prior unit test actually converts an xlsx file THROUGH the CLI
        # (only the poisoned-import subprocess test below touches xlsx at
        # all, and that one fails before ever reaching _convert_fn's xlsx
        # branch) — closes that gap directly.
        xlsx_path = _write_xlsx(tmp_path / "doc.xlsx")
        assert main([str(xlsx_path)]) == EXIT_OK
        assert capsys.readouterr().out != ""


class TestStdin:
    def test_requires_format(self) -> None:
        with pytest.raises(SystemExit) as exc:
            main([])
        assert exc.value.code == EXIT_USAGE

    def test_reads_and_converts(
        self, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
    ) -> None:
        data = build_minimal_docx(["From stdin"])
        monkeypatch.setattr(sys, "stdin", SimpleNamespace(buffer=io.BytesIO(data)))
        assert main(["--format", "docx"]) == EXIT_OK
        assert "From stdin" in capsys.readouterr().out

    def test_conversion_failure_prints_error_and_returns_typed_code(
        self, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
    ) -> None:
        monkeypatch.setattr(sys, "stdin", SimpleNamespace(buffer=io.BytesIO(b"not a zip at all")))
        assert main(["--format", "docx"]) == EXIT_CORRUPT_ARCHIVE
        assert "error:" in capsys.readouterr().err


class TestTypedExitCodes:
    def test_unsupported_format(self, tmp_path: Path) -> None:
        bad = tmp_path / "doc.docx"
        with zipfile.ZipFile(bad, "w") as z:
            z.writestr("hello.txt", "hi")
        assert main([str(bad)]) == EXIT_UNSUPPORTED_FORMAT

    def test_corrupt_archive(self, tmp_path: Path) -> None:
        bad = tmp_path / "doc.docx"
        bad.write_bytes(b"not a zip at all")
        assert main([str(bad)]) == EXIT_CORRUPT_ARCHIVE

    def test_missing_optional_dependency_via_subprocess(self, tmp_path: Path) -> None:
        # sys.modules poisoning only prevents the FIRST import of openpyxl in
        # a process — needs a fresh subprocess, same technique and reasoning
        # as test_optional_dependency_guards.py.
        xlsx_path = _write_xlsx(tmp_path / "doc.xlsx")
        script = (
            "import sys\n"
            "sys.modules['openpyxl'] = None\n"
            "from refigure.cli import main\n"
            f"sys.exit(main([{str(xlsx_path)!r}]))\n"
        )
        result = subprocess.run(
            [sys.executable, "-c", script],
            capture_output=True,
            text=True,
            cwd=str(REPO_ROOT),
            timeout=30,
        )
        assert result.returncode == EXIT_MISSING_DEPENDENCY
        assert "refigure[xlsx]" in result.stderr

    def test_exit_code_for_missing_optional_dependency_direct(self) -> None:
        # Fast, in-process complement to the subprocess test above: the
        # subprocess call above is invisible to coverage.py (subprocess
        # boundary), so this covers _exit_code_for's actual mapping line
        # directly rather than relying on that end-to-end test alone.
        from refigure.api import MissingOptionalDependencyError
        from refigure.cli import _exit_code_for

        assert _exit_code_for(MissingOptionalDependencyError("x")) == EXIT_MISSING_DEPENDENCY

    def test_unexpected_exception_is_internal_error(
        self,
        monkeypatch: pytest.MonkeyPatch,
        tmp_path: Path,
        capsys: pytest.CaptureFixture[str],
    ) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["hi"])

        import refigure.docx as docx_module

        def _boom(*args: object, **kwargs: object) -> None:
            raise RuntimeError("boom")

        monkeypatch.setattr(docx_module, "convert", _boom)
        assert main([str(doc)]) == EXIT_INTERNAL_ERROR
        assert "internal error" in capsys.readouterr().err

    def test_vlm_marker_limit_exceeded_is_its_own_exit_code(
        self, monkeypatch: pytest.MonkeyPatch, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        # Same technique as test_unexpected_exception_is_internal_error
        # above (mock docx.convert itself) — constructing a REAL document
        # that naturally triggers this exception needs a referenced VLM
        # marker above DOCX_IMAGE_MIN_BYTES, which none of this file's
        # existing docx-building helpers produce; the engine-level
        # behavior (pre-flight counting, judge-backfill counting) is
        # already exhaustively covered by tests/unit/test_vlm_max_markers.py
        # — this test only proves the CLI maps the exception to exit code
        # 7, not generic EXIT_INTERNAL_ERROR.
        doc = _write_docx(tmp_path / "doc.docx", ["hi"])

        import refigure.docx as docx_module
        from refigure.api import VlmMarkerLimitExceededError

        def _over_limit(*args: object, **kwargs: object) -> None:
            raise VlmMarkerLimitExceededError("doc.docx: 3 marker(s) exceed vlm_max_markers=2")

        monkeypatch.setattr(docx_module, "convert", _over_limit)
        assert main([str(doc), "--vlm", "--vlm-max-markers", "2"]) == EXIT_VLM_MARKER_LIMIT
        err = capsys.readouterr().err
        assert "internal error" not in err
        assert "vlm_max_markers=2" in err

    def test_exit_code_for_vlm_marker_limit_exceeded_direct(self) -> None:
        from refigure.api import VlmMarkerLimitExceededError
        from refigure.cli import _exit_code_for

        assert _exit_code_for(VlmMarkerLimitExceededError("x")) == EXIT_VLM_MARKER_LIMIT


class TestBatchMode:
    def test_requires_output_dir(self, tmp_path: Path) -> None:
        a = _write_docx(tmp_path / "a.docx", ["A"])
        b = _write_docx(tmp_path / "b.docx", ["B"])
        with pytest.raises(SystemExit) as exc:
            main([str(a), str(b)])
        assert exc.value.code == EXIT_USAGE

    def test_single_directory_triggers_batch(self, tmp_path: Path) -> None:
        _write_docx(tmp_path / "a.docx", ["A"])
        out_dir = tmp_path / "out"
        assert main([str(tmp_path), "-o", str(out_dir)]) == EXIT_OK
        assert (out_dir / "a.md").exists()

    def test_keep_going_partial_failure_summary_and_exit1(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        good = _write_docx(tmp_path / "good.docx", ["Good"])
        bad = tmp_path / "bad.docx"
        bad.write_bytes(b"not a zip at all")
        out_dir = tmp_path / "out"

        code = main([str(good), str(bad), "-o", str(out_dir)])

        assert code == EXIT_BATCH_PARTIAL_FAILURE
        assert (out_dir / "good.md").exists()
        assert not (out_dir / "bad.md").exists()
        assert "1/2 converted, 1 failed" in capsys.readouterr().err

    def test_fail_fast_aborts_before_writing_anything(self, tmp_path: Path) -> None:
        bad = tmp_path / "bad.docx"
        bad.write_bytes(b"not a zip at all")
        good = _write_docx(tmp_path / "good.docx", ["Good"])
        out_dir = tmp_path / "out"

        code = main([str(bad), str(good), "-o", str(out_dir), "--fail-fast"])

        assert code == EXIT_CORRUPT_ARCHIVE
        assert not out_dir.exists() or not any(out_dir.iterdir())

    def test_preserves_relative_structure_no_collision(self, tmp_path: Path) -> None:
        # anti-Docling-#3811 regression: same basename under 2 different
        # subdirectories of one walked root must not collide.
        _write_docx(tmp_path / "a" / "x.docx", ["A version"])
        _write_docx(tmp_path / "b" / "x.docx", ["B version"])
        out_dir = tmp_path / "out"

        assert main([str(tmp_path), "-o", str(out_dir)]) == EXIT_OK

        assert "A version" in (out_dir / "a" / "x.md").read_text(encoding="utf-8")
        assert "B version" in (out_dir / "b" / "x.md").read_text(encoding="utf-8")

    def test_explicit_collision_is_usage_error_and_writes_nothing(self, tmp_path: Path) -> None:
        first = _write_docx(tmp_path / "a" / "x.docx", ["A"])
        second = _write_docx(tmp_path / "b" / "x.docx", ["B"])
        out_dir = tmp_path / "out"

        code = main([str(first), str(second), "-o", str(out_dir)])

        assert code == EXIT_USAGE
        assert not out_dir.exists() or not any(out_dir.rglob("*.md"))

    def test_unrecognized_extension_among_batch_sources_is_usage_error(
        self, tmp_path: Path
    ) -> None:
        good = _write_docx(tmp_path / "good.docx", ["A"])
        stray = tmp_path / "notes.txt"
        stray.write_text("hi", encoding="utf-8")
        with pytest.raises(SystemExit) as exc:
            main([str(good), str(stray)])
        assert exc.value.code == EXIT_USAGE

    def test_nonexistent_source_among_batch_sources_is_usage_error(self, tmp_path: Path) -> None:
        good = _write_docx(tmp_path / "good.docx", ["A"])
        with pytest.raises(SystemExit) as exc:
            main([str(good), str(tmp_path / "nope.docx")])
        assert exc.value.code == EXIT_USAGE

    def test_empty_directory_source_is_usage_error(self, tmp_path: Path) -> None:
        # tmp_path itself has nothing convertible in it yet — a single
        # directory source is already batch mode (is_dir()), and an empty
        # plan must be rejected rather than silently writing 0 files.
        out_dir = tmp_path / "out"
        with pytest.raises(SystemExit) as exc:
            main([str(tmp_path), "-o", str(out_dir)])
        assert exc.value.code == EXIT_USAGE

    def test_duplicate_source_path_is_deduplicated(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        out_dir = tmp_path / "out"
        code = main([str(doc), str(doc), "-o", str(out_dir)])
        assert code == EXIT_OK
        assert "1/1 converted, 0 failed" in capsys.readouterr().err


class TestPlanBatchSymlinks:
    def test_symlinked_file_in_directory_source_is_excluded(self, tmp_path: Path) -> None:
        # Security-audit finding #15: following symlinks in a batch
        # directory walk would silently read from the symlink's real
        # target while the output path only reflects the symlink's own
        # apparent location — a read-side info-disclosure risk if refigure
        # is ever run over a directory tree an untrusted party could plant
        # symlinks into. Confirm the safe default (skip symlinked files,
        # no --follow-symlinks opt-in exists) while a genuine
        # same-directory file is still included.
        src_dir = tmp_path / "batch"
        real_file = _write_docx(tmp_path / "elsewhere" / "real.docx", ["Real content"])
        genuine = _write_docx(src_dir / "genuine.docx", ["Genuine content"])
        symlink = src_dir / "linked.docx"
        symlink.symlink_to(real_file)

        plan = _plan_batch([src_dir])

        input_paths = {file_path for file_path, _rel in plan}
        assert genuine in input_paths
        assert symlink not in input_paths
        assert len(plan) == 1


class TestJsonFlag:
    def test_single_file_is_valid_json_with_expected_keys(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        assert main([str(doc), "--json"]) == EXIT_OK
        payload = json.loads(capsys.readouterr().out)
        assert set(payload) == {
            "markdown",
            "warnings",
            "charts_found",
            "charts_rendered",
            "groups_found",
            "vlm_used",
        }
        assert "Hello" in payload["markdown"]

    def test_batch_writes_dot_json_not_dot_md(self, tmp_path: Path) -> None:
        _write_docx(tmp_path / "doc.docx", ["Hello"])
        out_dir = tmp_path / "out"
        assert main([str(tmp_path), "-o", str(out_dir), "--json"]) == EXIT_OK
        assert (out_dir / "doc.json").exists()
        assert not (out_dir / "doc.md").exists()
        json.loads((out_dir / "doc.json").read_text(encoding="utf-8"))  # parses cleanly


class TestVerbosityFlags:
    def test_default_shows_warnings_on_stderr(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        empty_doc = _write_docx(tmp_path / "empty.docx", [])
        assert main([str(empty_doc)]) == EXIT_OK
        assert "warning: no extractable content" in capsys.readouterr().err

    def test_batch_default_shows_per_file_warnings_on_stderr(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        empty_doc = _write_docx(tmp_path / "empty.docx", [])
        out_dir = tmp_path / "out"
        assert main([str(tmp_path), "-o", str(out_dir)]) == EXIT_OK
        assert f"warning: {empty_doc}: no extractable content" in capsys.readouterr().err

    def test_quiet_suppresses_warnings_not_stdout(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        empty_doc = _write_docx(tmp_path / "empty.docx", [])
        assert main([str(empty_doc), "-q"]) == EXIT_OK
        captured = capsys.readouterr()
        assert "warning:" not in captured.err

    def test_batch_quiet_still_prints_summary(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        _write_docx(tmp_path / "empty.docx", [])
        out_dir = tmp_path / "out"
        assert main([str(tmp_path), "-o", str(out_dir), "-q"]) == EXIT_OK
        captured = capsys.readouterr()
        assert "warning:" not in captured.err
        assert "1/1 converted, 0 failed" in captured.err

    def test_verbose_adds_per_file_batch_progress(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        _write_docx(tmp_path / "a.docx", ["A"])
        out_dir = tmp_path / "out"
        assert main([str(tmp_path), "-o", str(out_dir), "-v"]) == EXIT_OK
        assert "converted:" in capsys.readouterr().err

    def test_default_batch_has_no_per_file_progress(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        _write_docx(tmp_path / "a.docx", ["A"])
        out_dir = tmp_path / "out"
        assert main([str(tmp_path), "-o", str(out_dir)]) == EXIT_OK
        assert "converted:" not in capsys.readouterr().err

    def test_verbose_and_quiet_are_mutually_exclusive(self, tmp_path: Path) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["A"])
        with pytest.raises(SystemExit) as exc:
            main([str(doc), "-v", "-q"])
        assert exc.value.code == EXIT_USAGE


class TestMiscFlags:
    def test_version_prints_and_exits_zero(self, capsys: pytest.CaptureFixture[str]) -> None:
        with pytest.raises(SystemExit) as exc:
            main(["--version"])
        assert exc.value.code == 0
        assert "refigure" in capsys.readouterr().out

    def test_strict_flag_is_accepted_and_does_not_crash(self, tmp_path: Path) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        assert main([str(doc), "--strict"]) == EXIT_OK


class TestVlmFlags:
    """CLI wiring for --vlm and friends. The VLM ENGINE itself
    (Config.strict + soffice, enhance_docx_markdown's whole behavior) is
    already covered by tests/unit/vlm/test_vlm.py — these tests only prove
    the CLI builds the right Config/vlm_client, not that the engine works,
    same division of labor the spec calls for."""

    def _capture_config(self, monkeypatch: pytest.MonkeyPatch) -> dict[str, object]:
        captured: dict[str, object] = {}

        def _fake_convert_fn(fmt: str) -> object:
            def _convert(source: object, *, config: object = None) -> ConversionResult:
                captured["config"] = config
                return ConversionResult(markdown="ok")

            return _convert

        monkeypatch.setattr(cli_module, "_convert_fn", _fake_convert_fn)
        return captured

    def test_vlm_flag_sets_use_vlm_true(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)
        assert main([str(doc), "--vlm"]) == EXIT_OK
        assert captured["config"].use_vlm is True  # type: ignore[attr-defined]

    def test_no_vlm_flag_leaves_use_vlm_false(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)
        assert main([str(doc)]) == EXIT_OK
        assert captured["config"].use_vlm is False  # type: ignore[attr-defined]

    def test_vlm_verify_flag_sets_vlm_verify_true(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)
        assert main([str(doc), "--vlm", "--vlm-verify"]) == EXIT_OK
        assert captured["config"].vlm_verify is True  # type: ignore[attr-defined]

    def test_vlm_model_flag_overrides_config_default(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)
        assert main([str(doc), "--vlm", "--vlm-model", "some/model"]) == EXIT_OK
        assert captured["config"].vlm_model == "some/model"  # type: ignore[attr-defined]

    def test_vlm_model_unset_keeps_config_default(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from refigure.api import Config

        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)
        assert main([str(doc), "--vlm"]) == EXIT_OK
        assert captured["config"].vlm_model == Config().vlm_model  # type: ignore[attr-defined]

    def test_vlm_judge_mode_flag_overrides_config_default(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)
        assert main([str(doc), "--vlm", "--vlm-judge-mode", "solo"]) == EXIT_OK
        assert captured["config"].vlm_judge_mode == "solo"  # type: ignore[attr-defined]

    def test_vlm_max_markers_flag_overrides_config_default(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)
        assert main([str(doc), "--vlm", "--vlm-max-markers", "5"]) == EXIT_OK
        assert captured["config"].vlm_max_markers == 5  # type: ignore[attr-defined]

    def test_vlm_max_markers_unset_keeps_config_default(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from refigure.api import Config

        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)
        assert main([str(doc), "--vlm"]) == EXIT_OK
        assert captured["config"].vlm_max_markers == Config().vlm_max_markers  # type: ignore[attr-defined]

    def test_vlm_provider_openai_constructs_openai_client(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from refigure.vlm import client as vlm_client_module

        class _FakeOpenAIClient:
            def __init__(self, **kwargs: object) -> None:
                self.kwargs = kwargs

        monkeypatch.setattr(vlm_client_module, "OpenAIClient", _FakeOpenAIClient)
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)

        assert main([str(doc), "--vlm", "--vlm-provider", "openai", "--vlm-model", "m"]) == EXIT_OK

        vlm_client = captured["config"].vlm_client  # type: ignore[attr-defined]
        assert isinstance(vlm_client, _FakeOpenAIClient)
        assert vlm_client.kwargs["base_url"] is None
        assert vlm_client.kwargs["image_content_format"] == "dict"

    def test_vlm_provider_anthropic_constructs_anthropic_client(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from refigure.vlm import client as vlm_client_module

        class _FakeAnthropicClient:
            def __init__(self, **kwargs: object) -> None:
                self.kwargs = kwargs

        monkeypatch.setattr(vlm_client_module, "AnthropicClient", _FakeAnthropicClient)
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)

        assert (
            main([str(doc), "--vlm", "--vlm-provider", "anthropic", "--vlm-model", "m"]) == EXIT_OK
        )

        vlm_client = captured["config"].vlm_client  # type: ignore[attr-defined]
        assert isinstance(vlm_client, _FakeAnthropicClient)

    def test_vlm_model_missing_with_non_openrouter_provider_is_usage_error(
        self, tmp_path: Path
    ) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        with pytest.raises(SystemExit) as exc:
            main([str(doc), "--vlm", "--vlm-provider", "openai"])
        assert exc.value.code == EXIT_USAGE

    def test_vlm_base_url_without_openai_provider_is_usage_error(self, tmp_path: Path) -> None:
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        with pytest.raises(SystemExit) as exc:
            main([str(doc), "--vlm", "--vlm-base-url", "http://localhost:11434/v1/"])
        assert exc.value.code == EXIT_USAGE

    def test_vlm_api_key_file_content_becomes_client_api_key(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from refigure.vlm import client as vlm_client_module

        class _FakeOpenAIClient:
            def __init__(self, **kwargs: object) -> None:
                self.kwargs = kwargs

        monkeypatch.setattr(vlm_client_module, "OpenAIClient", _FakeOpenAIClient)
        key_file = tmp_path / "key.txt"
        key_file.write_text("  sk-from-file-content  \n", encoding="utf-8")
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)

        assert (
            main(
                [
                    str(doc),
                    "--vlm",
                    "--vlm-provider",
                    "openai",
                    "--vlm-model",
                    "m",
                    "--vlm-api-key-file",
                    str(key_file),
                ]
            )
            == EXIT_OK
        )

        vlm_client = captured["config"].vlm_client  # type: ignore[attr-defined]
        assert vlm_client.kwargs["api_key"] == "sk-from-file-content"

    def test_vlm_api_key_file_openrouter_becomes_config_vlm_api_key(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        key_file = tmp_path / "key.txt"
        key_file.write_text("sk-openrouter-key", encoding="utf-8")
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        captured = self._capture_config(monkeypatch)

        assert main([str(doc), "--vlm", "--vlm-api-key-file", str(key_file)]) == EXIT_OK

        config = captured["config"]
        assert config.vlm_client is None  # type: ignore[attr-defined]
        assert config.vlm_api_key == "sk-openrouter-key"  # type: ignore[attr-defined]

    def test_vlm_provider_openai_missing_credentials_is_typed_error_not_a_crash(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
    ) -> None:
        # Regression: _resolve_vlm_client constructs a REAL openai.OpenAI(...)
        # client eagerly, before any document is read — its own SDK raises
        # OpenAIError (not one of refigure's typed exceptions) when no
        # api_key/OPENAI_API_KEY is available. main() must translate this
        # into a clean exit code, not let it propagate as an unhandled
        # traceback.
        monkeypatch.delenv("OPENAI_API_KEY", raising=False)
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])

        code = main([str(doc), "--vlm", "--vlm-provider", "openai", "--vlm-model", "m"])

        assert code == EXIT_INTERNAL_ERROR
        assert "error:" in capsys.readouterr().err

    def test_vlm_flag_without_vlm_extra_is_missing_dependency_via_subprocess(
        self, tmp_path: Path
    ) -> None:
        # Same subprocess-poisoning technique as
        # TestTypedExitCodes.test_missing_optional_dependency_via_subprocess
        # — sys.modules poisoning only prevents the FIRST import of
        # pdfplumber in a process.
        doc = _write_docx(tmp_path / "doc.docx", ["Hello"])
        script = (
            "import sys\n"
            "sys.modules['pdfplumber'] = None\n"
            "from refigure.cli import main\n"
            f"sys.exit(main(['--vlm', {str(doc)!r}]))\n"
        )
        result = subprocess.run(
            [sys.executable, "-c", script],
            capture_output=True,
            text=True,
            cwd=str(REPO_ROOT),
            timeout=30,
        )
        assert result.returncode == EXIT_MISSING_DEPENDENCY
        assert "refigure[vlm]" in result.stderr

    def test_vlm_flag_on_xlsx_source_warns_and_still_converts(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        xlsx_path = _write_xlsx(tmp_path / "doc.xlsx")
        assert main([str(xlsx_path), "--vlm"]) == EXIT_OK
        assert "--vlm has no effect on .xlsx sources" in capsys.readouterr().err

    def test_vlm_flag_on_xlsx_source_in_batch_warns_once_per_file(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        _write_xlsx(tmp_path / "a.xlsx")
        _write_xlsx(tmp_path / "b.xlsx")
        out_dir = tmp_path / "out"
        assert main([str(tmp_path), "-o", str(out_dir), "--vlm"]) == EXIT_OK
        err = capsys.readouterr().err
        assert err.count("--vlm has no effect on .xlsx sources") == 2

    def test_vlm_flag_on_xlsx_source_via_stdin_warns(
        self,
        monkeypatch: pytest.MonkeyPatch,
        capsys: pytest.CaptureFixture[str],
    ) -> None:
        buf = io.BytesIO()
        wb = openpyxl.Workbook()
        assert wb.active is not None
        wb.active.append(["a"])
        wb.save(buf)
        monkeypatch.setattr(sys, "stdin", SimpleNamespace(buffer=io.BytesIO(buf.getvalue())))
        assert main(["--format", "xlsx", "--vlm"]) == EXIT_OK
        assert "--vlm has no effect on .xlsx sources" in capsys.readouterr().err


class TestModuleEntrypoint:
    def test_main_module_is_importable(self) -> None:
        # refigure.__main__ is otherwise never imported in-process (only
        # via the subprocess test below, a separate process coverage.py
        # can't see) — this covers its own top-level statements directly;
        # the `if __name__ == "__main__":` guard stays pragma-excluded,
        # see refigure/__main__.py.
        import refigure.__main__  # noqa: F401

    def test_python_dash_m_refigure_help_exits_zero(self) -> None:
        # Nothing else in this suite invokes the real `python -m refigure`
        # entrypoint (refigure/__main__.py) — everything else calls
        # cli.main() in-process. A subprocess is required here for the
        # same reason as test_optional_dependency_guards.py: it's the only
        # way to actually exercise `if __name__ == "__main__":`.
        result = subprocess.run(
            [sys.executable, "-m", "refigure", "--help"],
            capture_output=True,
            text=True,
            cwd=str(REPO_ROOT),
            timeout=30,
        )
        assert result.returncode == 0
        assert "usage:" in result.stdout
