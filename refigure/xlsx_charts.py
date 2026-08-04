"""Embedded xlsx charts (spec convert-xlsx §3): the architectural analog of
``docx_groups.py``, adapted to the streamless (cell-anchored) model — an
xlsx chart has no "position within a paragraph", only a sheet + anchor cell,
so instead of an in-stream sentinel the marker is declaratively placed right
after its sheet's table (``converters._convert_xlsx``). Hence the different
name for the detection function too: unlike the docx analog
(``extract_and_strip_groups``, which rewrites ``word/document.xml`` to cut
the group out), here ``raw`` is NOT modified at all — calling this "strip"
would be inaccurate.

Detection: ``xl/worksheets/sheetK.xml`` carries ``<drawing r:id="rIdX"/>`` ->
``xl/worksheets/_rels/sheetK.xml.rels`` -> ``xl/drawings/drawingM.xml`` ->
``<xdr:oneCellAnchor>``/``<xdr:twoCellAnchor>`` with ``<xdr:from>`` (anchor
cell: 0-indexed col/row) and a ``graphicFrame`` carrying ``c:chart`` (a
reference to ``xl/charts/chartN.xml`` via
``xl/drawings/_rels/drawingM.xml.rels``). Sheet name -> path of its XML part
goes through ``xl/workbook.xml`` (name -> r:id) +
``xl/_rels/workbook.xml.rels`` (r:id -> Target); three different sources of
relative paths (workbook/sheet/drawing) each require resolving relative to
the directory of THAT SPECIFIC source part (not a hardcoded ``xl/`` as in
docx, where all rels live under a single ``word/`` — see
``_resolve_target``).

``id12`` is the sha256 of the chart's XML STRUCTURE (``etree.tostring`` of
the parsed ``xl/charts/chartN.xml``), NOT of rendered bytes: the same
principle used for docx groups (the one deviation that docx tests tripped
over twice — recorded here in the docstring up front as a heads-up).
``captions`` come from the chart part's ``c:title``: the identical DrawingML
schema ``c:tx/c:rich/a:p/a:r/a:t`` used by native docx charts (§2-ter
convert-docx) — the code is adapted to the different XML paths, not copied
1:1.
"""

from __future__ import annotations

import hashlib
import io
import posixpath
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from lxml import etree
from openpyxl.utils import get_column_letter

_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "c": "http://schemas.openxmlformats.org/drawingml/2006/chart",
}
# posOffset/ext coordinates that show up in itertext(), see docx_groups
_NUMERIC_JUNK_RE = re.compile(r"^-?\d+$")


def _q(prefix: str, local: str) -> str:
    return f"{{{_NS[prefix]}}}{local}"


@dataclass(frozen=True)
class XlsxChart:
    id12: str
    sheet: str
    anchor_cell: str
    captions: tuple[str, ...]


def _filter_caption_texts(texts: Any) -> tuple[str, ...]:
    """Same filter as ``docx_groups._filter_caption_texts`` — discards numeric
    geometry junk and deduplicates in order of first appearance."""
    seen: set[str] = set()
    out: list[str] = []
    for t in texts:
        s = t.strip()
        if not s or _NUMERIC_JUNK_RE.match(s) or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return tuple(out)


def _rel_targets(z: zipfile.ZipFile, part: str) -> dict[str, str]:
    """rId -> raw Target (not yet resolved) for ``part``, read from the adjacent
    .rels."""
    rels_name = f"{posixpath.dirname(part)}/_rels/{posixpath.basename(part)}.rels"
    if rels_name not in z.namelist():
        return {}
    root = etree.fromstring(z.read(rels_name))
    return {rel.get("Id"): rel.get("Target") for rel in root if rel.get("Id")}


def _resolve_target(source_part: str, target: str) -> str:
    """OPC-resolve a Target relative to its OWN source part: absolute (leading
    ``/``) -> package root, otherwise -> relative to source_part's directory
    (not a single hardcoded directory — unlike docx, where all rels live
    under ``word/``; here the chain involves ``xl/``, ``xl/worksheets/``,
    and ``xl/drawings/``)."""
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


def _sheet_parts(z: zipfile.ZipFile) -> dict[str, str]:
    """Sheet name (as in ``wb.sheetnames``) -> path of the sheet's XML part in the
    zip."""
    names = set(z.namelist())
    if "xl/workbook.xml" not in names:
        return {}
    rel_targets = _rel_targets(z, "xl/workbook.xml")
    root = etree.fromstring(z.read("xl/workbook.xml"))
    out: dict[str, str] = {}
    for sheet_el in root.findall(f".//{_q('main', 'sheet')}"):
        name, rid = sheet_el.get("name"), sheet_el.get(_q("r", "id"))
        if name is None or rid is None or rid not in rel_targets:
            continue
        part = _resolve_target("xl/workbook.xml", rel_targets[rid])
        if part in names:
            out[name] = part
    return out


def _chart_anchors(drawing_root: Any) -> list[tuple[Any, Any]]:
    """(anchor, chart_ref) for every ``c:chart`` inside every
    ``xdr:oneCellAnchor``/``xdr:twoCellAnchor`` — usually a lone
    ``graphicFrame`` (one chart per anchor), but an anchor can legitimately
    be an ``xdr:grpSp`` (Excel's "group these charts" feature) wrapping
    SEVERAL charts at one shared position: real, confirmed structure (e.g.
    eia-aeo-2026-figures.xlsx's drawing25.xml, one anchor nesting 6 charts
    under one grpSp, each in its own graphicFrame) — every one must be
    yielded, not just the first. A prior version used a singular ``.find()``
    here, which silently dropped every chart after the first in a grouped
    anchor (found via corpus testing, stage 5, 2026-08-05: 5/43, 48/93, and
    1/22 charts silently lost across 3 real fixtures, with no trace in
    ``ConversionResult`` at all — the exact opposite of this project's
    zero-loss positioning)."""
    pairs: list[tuple[Any, Any]] = []
    for tag in ("oneCellAnchor", "twoCellAnchor"):
        for anchor in drawing_root.findall(_q("xdr", tag)):
            for chart_ref in anchor.findall(f".//{_q('c', 'chart')}"):
                pairs.append((anchor, chart_ref))
    return pairs


def _anchor_col_row(anchor: Any) -> tuple[int, int]:
    """``xdr:from`` (0-indexed col/row) -> (col, row) 1-indexed, as expected by
    ``openpyxl.utils.column_index_from_string``/``ws.cell``."""
    frm = anchor.find(_q("xdr", "from"))
    return int(frm.findtext(_q("xdr", "col"))) + 1, int(frm.findtext(_q("xdr", "row"))) + 1


def _anchor_cell(anchor: Any) -> str:
    """``xdr:from`` (0-indexed col/row) -> an Excel cell reference (``D2``)."""
    col, row = _anchor_col_row(anchor)
    return f"{get_column_letter(col)}{row}"


def _chart_title(chart_root: Any) -> tuple[str, ...]:
    title = chart_root.find(f".//{_q('c', 'title')}")
    if title is None:
        return ()
    return _filter_caption_texts(title.itertext())


def iter_chart_entries(raw: Path | bytes) -> list[tuple[XlsxChart, Any]]:
    """A single pass over the workbook, sheet by sheet: (metadata, parsed
    ``chart_root``) for every chart — one single pass over the zip/XML.
    Public (no ``_`` prefix) because a consumer that needs BOTH slices at
    once (``converters._convert_xlsx`` — metadata for grouping by sheet /
    sorting by anchor, AND roots for ``chart_data.parse_chart``) must call
    it directly exactly ONCE — otherwise (a real bug found in review)
    ``extract_charts(raw)`` + ``extract_chart_roots(raw)`` used together
    would read and parse the workbook TWICE. ``extract_charts``/
    ``extract_chart_roots`` below are thin wrappers for consumers that need
    only one slice (tests, one-off checks). ``raw`` is NOT modified (see the
    module docstring). A malformed/unreachable reference at any step of the
    chain (a sheet with no drawing, a drawing with no rels, a missing chart
    part) is silently skipped (a terminal safety net — conversion doesn't
    fail on corrupted OOXML, symmetric with ``_classify_docx``).

    ``raw`` may be ``bytes`` (refigure accepts in-memory input, §2
    stage2-public-api-wrapper) — ``zipfile.ZipFile`` doesn't read raw
    ``bytes`` directly (only a path or a file-like object), so we wrap it in
    ``io.BytesIO``."""
    source = raw if isinstance(raw, Path) else io.BytesIO(raw)
    with zipfile.ZipFile(source) as z:
        names = set(z.namelist())
        entries: list[tuple[XlsxChart, Any]] = []
        for sheet_name, sheet_part in _sheet_parts(z).items():
            if sheet_part not in names:
                continue
            sheet_rels = _rel_targets(z, sheet_part)
            drawing_ref = etree.fromstring(z.read(sheet_part)).find(_q("main", "drawing"))
            if drawing_ref is None:
                continue
            drid = drawing_ref.get(_q("r", "id"))
            if drid is None or drid not in sheet_rels:
                continue
            drawing_part = _resolve_target(sheet_part, sheet_rels[drid])
            if drawing_part not in names:
                continue
            drawing_rels = _rel_targets(z, drawing_part)
            drawing_root = etree.fromstring(z.read(drawing_part))
            for anchor, chart_ref in _chart_anchors(drawing_root):
                crid = chart_ref.get(_q("r", "id"))
                if crid is None or crid not in drawing_rels:
                    continue
                chart_part = _resolve_target(drawing_part, drawing_rels[crid])
                if chart_part not in names:
                    continue
                chart_root = etree.fromstring(z.read(chart_part))
                entries.append(
                    (
                        XlsxChart(
                            id12=hashlib.sha256(etree.tostring(chart_root)).hexdigest()[:12],
                            sheet=sheet_name,
                            anchor_cell=_anchor_cell(anchor),
                            captions=_chart_title(chart_root),
                        ),
                        chart_root,
                    )
                )
        return entries


def extract_charts(raw: Path) -> list[XlsxChart]:
    """All embedded charts of the workbook, sheet by sheet (metadata —
    id12/sheet/anchor/captions, WITHOUT the parsed XML). A convenience
    wrapper around ``iter_chart_entries`` for consumers that don't need the
    roots (e.g. a one-off check in a test) — for using metadata AND roots
    together, see the ``iter_chart_entries`` docstring."""
    return [entry for entry, _root in iter_chart_entries(raw)]


def extract_chart_roots(raw: Path) -> dict[str, Any]:
    """id12 -> parsed ``chart_root`` (spec chart-data-extraction §4.1) — the
    input for ``chart_data.parse_chart``. A convenience wrapper around
    ``iter_chart_entries`` for consumers that don't need the metadata
    separately (e.g. a one-off check in a test)."""
    return {entry.id12: root for entry, root in iter_chart_entries(raw)}


def _chart_refs(chart_root: Any) -> list[tuple[str, str]]:
    """(sheet_name, cell_range_text) for every ``<c:f>`` series formula of the
    chart (unquoting a sheet name containing spaces: ``'My Sheet'!$A$1`` ->
    ``My Sheet``, ``''``->``'`` — standard Excel apostrophe escaping)."""
    out: list[tuple[str, str]] = []
    for f in chart_root.findall(f".//{_q('c', 'f')}"):
        text = f.text or ""
        if "!" not in text:
            continue
        sheet, _, cell_range = text.partition("!")
        sheet = sheet.strip()
        if sheet.startswith("'") and sheet.endswith("'"):
            sheet = sheet[1:-1].replace("''", "'")
        if sheet:
            out.append((sheet, cell_range))
    return out


def render_chart_marker(chart: XlsxChart) -> str:
    # English literal — see docx_groups._render_group_marker (B17, spec
    # convert-knowledge-seam-hardening §1): the marker text gets indexed.
    caption_line = "; ".join(chart.captions) if chart.captions else "(no captions)"
    return (
        f"> [Figure, xlsx chart {chart.id12} on {chart.sheet}!{chart.anchor_cell} — "
        f"chart content not analyzed]\n"
        f"> captions: {caption_line}"
    )
