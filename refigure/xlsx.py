"""Public XLSX -> Markdown conversion entry point.

Imports openpyxl at module level — nothing else in this package touches it,
so ``pip install refigure`` (no extra) never pulls it in. Importing this
module without ``refigure[xlsx]`` installed raises
``MissingOptionalDependencyError`` immediately, with an actionable message,
instead of a bare ``ModuleNotFoundError``.
"""

from __future__ import annotations

import datetime as _dt
import io
import zipfile
from pathlib import Path
from typing import Any, BinaryIO

from . import chart_data, chart_render, xlsx_charts, zipsafe
from ._io import normalize_source
from .api import (
    Config,
    ConversionResult,
    CorruptArchiveError,
    MissingOptionalDependencyError,
    UnsupportedFormatError,
)

try:
    import openpyxl
    from openpyxl.utils.cell import coordinate_to_tuple
except ImportError as exc:
    raise MissingOptionalDependencyError(
        "refigure[xlsx] is required to convert XLSX files"
    ) from exc


def _xlsx_cell_str(value: Any) -> str:
    """Cell value -> GFM table cell text. Dates/datetimes -> ISO; an
    integer-valued float -> without ".0"; everything else -> str(). Newlines
    become spaces (pipe characters are not escaped — same accepted
    convention as elsewhere in this pipeline)."""
    if value is None:
        return ""
    if isinstance(value, (_dt.datetime, _dt.date)):
        text = value.isoformat()
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    return text.strip().replace("\n", " ")


def _sheet_is_empty(ws: Any) -> bool:
    return all(cell.value is None for row in ws.iter_rows() for cell in row)


def _sheet_table(ws: Any) -> str:
    """Sheet -> GFM table: the first row of the used range is the structural
    header (not semantic — mirrors already-accepted markdownify-table
    behavior for docx without <thead>)."""
    rows = [[_xlsx_cell_str(c.value) for c in row] for row in ws.iter_rows()]
    header, *body = rows
    lines = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * len(header)) + " |"]
    lines.extend("| " + " | ".join(row) + " |" for row in body)
    return "\n".join(lines)


def _render_xlsx_chart_block(chart: Any, chart_root: Any) -> tuple[str, bool]:
    """Data-driven render of one xlsx chart: parse_chart -> render_chart;
    empty extraction (no numCache) -> caption-fallback marker, same as
    before. The provenance line (sheet+anchor) is added here, not inside
    chart_render (container-agnostic) — it carries the positional context
    the old VLM marker used to ("on {sheet}!{anchor}")."""
    rendered = chart_render.render_chart(chart_data.parse_chart(chart_root))
    if rendered is None:
        return xlsx_charts.render_chart_marker(chart), False
    provenance = f"> sheet {chart.sheet}, anchor {chart.anchor_cell}"
    return f"{provenance}\n\n{rendered}", True


def convert(source: Path | bytes | BinaryIO, *, config: Config | None = None) -> ConversionResult:
    """Convert an XLSX file (path, bytes, or a file-like object) to Markdown."""
    config = config or Config()
    normalized = normalize_source(source)

    try:
        zipsafe.check_archive(normalized)
    except (zipsafe.ArchiveBombSuspected, zipfile.BadZipFile) as exc:
        raise CorruptArchiveError(str(exc)) from exc

    # openpyxl accepts a path or a file-like object, but not raw bytes
    # directly (verified live) — wrap only when needed.
    wb_source = normalized if isinstance(normalized, Path) else io.BytesIO(normalized)
    try:
        wb = openpyxl.load_workbook(wb_source, data_only=True, read_only=False)
    except KeyError as exc:
        # openpyxl's own signal for "valid zip, but not an xlsx" (verified
        # live: KeyError on the missing [Content_Types].xml part).
        raise UnsupportedFormatError(str(exc)) from exc

    chart_roots: dict[str, Any] = {}
    charts_by_sheet: dict[str, list[Any]] = {}
    for chart, chart_root in xlsx_charts.iter_chart_entries(normalized):
        chart_roots[chart.id12] = chart_root
        charts_by_sheet.setdefault(chart.sheet, []).append(chart)
    for charts in charts_by_sheet.values():
        charts.sort(key=lambda c: coordinate_to_tuple(c.anchor_cell))

    sections: list[str] = []
    any_content = False
    charts_found = 0
    charts_rendered = 0
    for name in wb.sheetnames:
        ws = wb[name]
        heading = f"## {name}" if ws.sheet_state == "visible" else f"## {name} (hidden)"
        sheet_charts = charts_by_sheet.get(name, [])
        if _sheet_is_empty(ws) and not sheet_charts:
            sections.append(f'{heading}\n\n> [Sheet "{name}" — empty, skipped]')
            continue
        any_content = True
        parts = [heading]
        if not _sheet_is_empty(ws):
            parts.append(_sheet_table(ws))
        if sheet_charts:
            blocks: list[str] = []
            for c in sheet_charts:
                block, rendered = _render_xlsx_chart_block(c, chart_roots[c.id12])
                blocks.append(block)
                charts_found += 1
                charts_rendered += rendered
            parts.append("\n\n".join(blocks))
        sections.append("\n\n".join(parts))

    warnings: list[str] = []
    if not any_content:
        warnings.append("no extractable content")
    if charts_found and not chart_render.mermaidx_available():
        warnings.append(
            "mermaidx not installed — chart diagrams disabled, tables only "
            "(install refigure[xlsx] with mermaidx to enable rendering)"
        )

    return ConversionResult(
        markdown="\n\n".join(sections) + "\n",
        warnings=warnings,
        charts_found=charts_found,
        charts_rendered=charts_rendered,
        groups_found=0,
        vlm_used=False,
    )
